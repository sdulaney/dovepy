#! /usr/bin/env python3

#Module is used to help assist in the abstraction of retrieving data in a manner specific to 
#the application from excel spreadsheets

import openpyxl
import os.path
import sys
import random

#Helper function to add data. This should not be called directly outside this file. The function
#takes the created exel workbook from append_outfile and populates col_loc with the data from data_array
#adding the column name header_name if not already existant. 
def __load_col(workbook, data_array, file_path, col_loc, header_name):
    wb = workbook
    sheet = wb.active

    print('\tAppending ' + header_name + ' to column ' + col_loc)

    sheet[col_loc + '1'] = header_name
    for i, data in enumerate(data_array):
        sheet[col_loc + str(i + 2)] = data

    return wb

#Load data back into the specified .xlsx at file_path
def load_data(file_path, col_loc, header_name, data):
    w = openpyxl.load_workbook(file_path)
    wb = __load_col(w, data, file_path, col_loc, header_name)
    wb.save(file_path)

#Search for the row containing the string
def search_rows(file_path, string, col_letter):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    for row_index in sheet[col_letter + '1:C' + str(sheet.max_row)]:
        for cell in row_index:
            if cell.value[0:10] == string[0:10]:
                return cell.row

#Creates an exel workbook if one does not already exist in which it will open the existing file at file_path
#col_loc and header_name are used to pass to the helper function __load_col which 
def create_initial_outfile(data_array, file_path, col_loc, header_name):
    if os.path.isfile(file_path):
        wb = openpyxl.load_workbook(file_path)
        wb = __load_col(wb, data_array, file_path, col_loc, header_name)
        wb.save(file_path)
    else:
        wb = openpyxl.Workbook()
        wb = __load_col(wb, data_array, file_path, col_loc, header_name)
        wb.save(file_path)

#Will return a column of data from the specified file path (must be .xls file) and returns the data
#in an array
def get_column_data(columnHeaderName, file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    max_row = str(sheet.max_row)
    row_start = '2'
    headers = __pair_header_with_col_ids(workbook)
    col_id = ''
    data = []

    for d in headers:
        if d['Header_Name'] == columnHeaderName:
            col_id = d['Col_ID']
        
    for rowObj in sheet[col_id + row_start : col_id + max_row]:
        for cell in rowObj:
            data.append(str(cell.value))
    return data

#Helper function to get_column_data that pairs a text arguement to the corresponding header column letter.
#This function should not be called outside this module. 
def __pair_header_with_col_ids(workbook):
  headerDict = []
  count = 1
  sheet = workbook.active

  for header in sheet['A1': 'P1']:
    for cellObj in header:
      if cellObj.value:
        headerDict.append({'Header_Name': cellObj.value, 'Col_ID': openpyxl.cell.get_column_letter(count)})
        count = count + 1
  return headerDict

#Prepare necessary in data for prospectnow scrape. This function takes 4 lists of data that must
#be of equal length and appends necessary ',' for each element in the 4 lists. The return is an
#address formated as such (1234 Main St., Gotham City, Gotham 12345) Max is 1000. Anything over
#1000 addresses will be truncated. Returns address list 
def build_address_list(address_list, city_list, state_list, zip_code_list):
    _list = []
    length = len(address_list)

    for i in range(len(address_list)):
        string = address_list[i] + ', ' + city_list[i] + ', ' + state_list[i] + ', ' + zip_code_list[i]
        _list.append(string)

    if length <= 1000:
        return_list = _list
        print('\nCreated ' + str(len(address_list)) + ' addresses')
        return return_list

    else:
        return_list = _list[0:1000]
        check = input('\nCreated ' + str(len(address_list)) + ' addresses. Only first 1000 will be used. Continue? y/n\n')
        
        if check == 'y':
            return return_list

        elif check == 'n':
            print('Exiting...')
            sys.exit()

#Write to one specific cell and save workbook. Must call search function first to work properly
def write_cell(file_path, row, col, string, header_name):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    if not sheet.cell(row=1, column=col).value:
        sheet.cell(row=1, column=col).value = header_name
    sheet.cell(row=row, column=col).value = string
    workbook.save(file_path)

#Return a list of dictionary objects containing the owner address parts
#Depreciated
def format_owner_address(file_path, address_col, city_col, state_col, zip_col):
    addresses = get_column_data(address_col, file_path)
    cities = get_column_data(city_col, file_path)
    states = get_column_data(state_col, file_path)
    zips = get_column_data(zip_col, file_path)

    dict_list = []
    for i, add in enumerate(addresses):
        d = {'address':addresses[i], 'city':cities[i], 'state':states[i], 'zip':zips[i]}
        dict_list.append(d)

    return dict_list

#Get property_key at specified row
def get_property_key(file_path, row):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    return sheet.cell(row=row, column=1).value


#Same as above but selects a random number from the list
def build_random_address_list(address_list, city_list, state_list, zip_code_list, num_to_grab):
    _list = []
    length = len(address_list)

    for i in range(len(address_list)):
        string = address_list[i] + ', ' + city_list[i] + ', ' + state_list[i] + ', ' + zip_code_list[i]
        _list.append(string)

    random.shuffle(_list)
    return _list[0:num_to_grab]





